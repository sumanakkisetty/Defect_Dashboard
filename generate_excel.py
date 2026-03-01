"""
ROCm Defect Tracker — Excel Dashboard Generator
Creates ROCm_Defect_Dashboard.xlsx with a self-contained dashboard.
No Python required to VIEW the file — just open in Excel.

Run: python generate_excel.py
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.chart.series import DataPoint
from datetime import datetime
import warnings
warnings.filterwarnings("ignore")

# ── Load & Prepare Data ────────────────────────────────────────────────────────
df = pd.read_csv("defects.csv", parse_dates=["Created_Date", "Last_Updated_Date"])
df["Week_Start"] = df["Created_Date"].dt.to_period("W").apply(lambda p: p.start_time)
df["Week_Label"] = df["Week_Start"].dt.strftime("%d %b")

PRIORITY_ORDER = ["Critical", "High", "Medium", "Low"]
SEVERITY_ORDER = ["S1 - Blocker", "S2 - Critical", "S3 - Major", "S4 - Minor"]
STATUS_ORDER   = ["Open", "In Progress", "In Review", "Resolved", "Closed", "Won't Fix"]
DIRECTOR_ORDER = ["SAM", "Kumar", "Philips", "Suma", "Ravi"]

DIRECTOR_TEAMS = {
    "SAM":    "Compiler, Runtime, Driver",
    "Kumar":  "Math Libraries",
    "Philips":"Tools & Infra",
    "Suma":   "Release Engineering",
    "Ravi":   "Documentation",
}

# ── Pre-compute all summaries ──────────────────────────────────────────────────
# WoW by Priority
wow_raw = (df.groupby(["Week_Start", "Week_Label", "Priority"])
             .size().unstack(fill_value=0).reset_index()
             .sort_values("Week_Start"))
for p in PRIORITY_ORDER:
    if p not in wow_raw.columns:
        wow_raw[p] = 0

# Director × Severity
dir_sev_raw = (df.groupby(["Director", "Severity"])
                 .size().unstack(fill_value=0)
                 .reindex(DIRECTOR_ORDER, fill_value=0)
                 .reset_index())
for s in SEVERITY_ORDER:
    if s not in dir_sev_raw.columns:
        dir_sev_raw[s] = 0

# Triage × Status
triage_raw = (df.groupby(["Triage_Assignment", "Status"])
                .size().unstack(fill_value=0).reset_index())
for s in STATUS_ORDER:
    if s not in triage_raw.columns:
        triage_raw[s] = 0

# Director summary table
dir_total = df.groupby("Director").size().rename("Total")
dir_open  = df[df["Status"] == "Open"].groupby("Director").size().rename("Open")
dir_crit  = df[df["Priority"] == "Critical"].groupby("Director").size().rename("Critical")
dir_res   = df[df["Status"].isin(["Resolved", "Closed"])].groupby("Director").size().rename("Resolved")
dir_sum   = pd.concat([dir_total, dir_open, dir_crit, dir_res], axis=1).fillna(0).astype(int)

# Count tables
priority_cnts  = df["Priority"].value_counts().reindex(PRIORITY_ORDER, fill_value=0)
severity_cnts  = df["Severity"].value_counts().reindex(SEVERITY_ORDER, fill_value=0)
status_cnts    = df["Status"].value_counts().reindex(STATUS_ORDER,   fill_value=0)
category_cnts  = df["Triage_Category"].value_counts().head(12)
open_df        = df[df["Status"].isin(["Open","In Progress","In Review"]) & (df["Assignee"] != "")]
top_asgn       = open_df["Assignee"].value_counts().head(10)
release_cnts   = df["Target_Release"].value_counts()

# WoW by Director (line chart)
wow_dir_raw = (df.groupby(["Week_Start", "Week_Label", "Director"])
                 .size().unstack(fill_value=0).reset_index()
                 .sort_values("Week_Start"))
for d in DIRECTOR_ORDER:
    if d not in wow_dir_raw.columns:
        wow_dir_raw[d] = 0

# KPIs
KPI_VALUES = {
    "Total Defects":   len(df),
    "Open":            int(status_cnts.get("Open", 0)),
    "S1 Blockers":     int(severity_cnts.get("S1 - Blocker", 0)),
    "Critical (P1)":   int(priority_cnts.get("Critical", 0)),
    "In Progress":     int(status_cnts.get("In Progress", 0)),
    "Resolved/Closed": int(df["Status"].isin(["Resolved","Closed"]).sum()),
}

print("Summaries calculated. Building workbook...")

# ── Style Helpers ──────────────────────────────────────────────────────────────
def fill(hex6):
    return PatternFill("solid", fgColor=hex6)

def font(bold=False, size=11, color="000000", italic=False, name="Calibri"):
    return Font(bold=bold, size=size, color=color, italic=italic, name=name)

def border(color="BBBBBB", style="thin"):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def align(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def set_cell(ws, row, col, value="", bold=False, size=11, fcolor="000000",
             bg=None, italic=False, h_align="center", wrap=False, bdr=None):
    c = ws.cell(row=row, column=col)
    c.value     = value
    c.font      = font(bold=bold, size=size, color=fcolor, italic=italic)
    c.alignment = align(h_align, wrap=wrap)
    if bg:
        c.fill = fill(bg)
    if bdr:
        c.border = bdr
    return c

# ── Color Palette ──────────────────────────────────────────────────────────────
NAVY    = "1B2631"
SLATE   = "2E4057"
PURPLE  = "4A235A"
RED     = "C0392B"
ORANGE  = "CA6F1E"
AMBER   = "9A7D0A"
GREEN   = "1A7A4A"
LGRAY   = "F2F3F4"
MGRAY   = "D5D8DC"
WHITE   = "FFFFFF"

KPI_STYLES = [
    ("TOTAL DEFECTS",   SLATE,  WHITE),
    ("OPEN",            RED,    WHITE),
    ("S1 BLOCKERS",     RED,    WHITE),
    ("CRITICAL (P1)",   ORANGE, WHITE),
    ("IN PROGRESS",     AMBER,  WHITE),
    ("RESOLVED/CLOSED", GREEN,  WHITE),
]

PCOL_HEX = {"Critical":"C0392B","High":"E67E22","Medium":"D4AC0D","Low":"27AE60"}
SCOL_HEX = {"S1 - Blocker":"C0392B","S2 - Critical":"E67E22",
            "S3 - Major":"D4AC0D","S4 - Minor":"27AE60"}
STCOL_HEX= {"Open":"C0392B","In Progress":"E67E22","In Review":"D4AC0D",
             "Resolved":"27AE60","Closed":"7F8C8D","Won't Fix":"8E44AD"}
DCOL_HEX = {"SAM":"4A235A","Kumar":"0B5345","Philips":"154360",
             "Suma":"784212","Ravi":"78281F"}

# ─────────────────────────────────────────────────────────────────────────────
# ── 1.  HIDDEN SUMMARY SHEETS  (data sources for charts) ──────────────────────
# ─────────────────────────────────────────────────────────────────────────────
wb = Workbook()
wb.remove(wb.active)

def write_hidden(name, headers, rows):
    ws = wb.create_sheet(name)
    ws.sheet_state = "hidden"
    for ci, h in enumerate(headers, 1):
        ws.cell(1, ci).value = h
    for ri, row in enumerate(rows, 2):
        for ci, v in enumerate(row, 1):
            ws.cell(ri, ci).value = v
    return ws, len(rows)

# WoW
wow_headers = ["Week"] + PRIORITY_ORDER
wow_rows    = [[r["Week_Label"]] + [int(r.get(p, 0)) for p in PRIORITY_ORDER]
               for _, r in wow_raw.iterrows()]
ws_wow, n_wow = write_hidden("_WoW", wow_headers, wow_rows)

# Director × Severity
dir_headers = ["Director"] + SEVERITY_ORDER
dir_rows    = [[r["Director"]] + [int(r.get(s, 0)) for s in SEVERITY_ORDER]
               for _, r in dir_sev_raw.iterrows()]
ws_dir, n_dir = write_hidden("_Dir", dir_headers, dir_rows)

# Triage × Status
tr_headers  = ["Team"] + STATUS_ORDER
tr_rows     = [[r["Triage_Assignment"]] + [int(r.get(s, 0)) for s in STATUS_ORDER]
               for _, r in triage_raw.iterrows()]
ws_tr, n_tr = write_hidden("_Triage", tr_headers, tr_rows)

# WoW by Director
wdir_headers = ["Week"] + DIRECTOR_ORDER
wdir_rows    = [[r["Week_Label"]] + [int(r.get(d, 0)) for d in DIRECTOR_ORDER]
                for _, r in wow_dir_raw.iterrows()]
ws_wdir, n_wdir = write_hidden("_WDir", wdir_headers, wdir_rows)

# Count tables (single sheet, side by side)
ws_c = wb.create_sheet("_Counts")
ws_c.sheet_state = "hidden"
# Col A-B: Priority
ws_c["A1"] = "Priority"; ws_c["B1"] = "Count"
for i, (k, v) in enumerate(priority_cnts.items(), 2):
    ws_c.cell(i, 1).value = k; ws_c.cell(i, 2).value = int(v)
# Col D-E: Severity
ws_c["D1"] = "Severity"; ws_c["E1"] = "Count"
for i, (k, v) in enumerate(severity_cnts.items(), 2):
    ws_c.cell(i, 4).value = k; ws_c.cell(i, 5).value = int(v)
# Col G-H: Status
ws_c["G1"] = "Status"; ws_c["H1"] = "Count"
for i, (k, v) in enumerate(status_cnts.items(), 2):
    ws_c.cell(i, 7).value = k; ws_c.cell(i, 8).value = int(v)
# Col J-K: Categories
ws_c["J1"] = "Category"; ws_c["K1"] = "Count"
for i, (k, v) in enumerate(category_cnts.items(), 2):
    ws_c.cell(i, 10).value = k; ws_c.cell(i, 11).value = int(v)
# Col M-N: Top Assignees
ws_c["M1"] = "Assignee"; ws_c["N1"] = "Open"
for i, (k, v) in enumerate(top_asgn.items(), 2):
    ws_c.cell(i, 13).value = k; ws_c.cell(i, 14).value = int(v)
# Col P-Q: Release
ws_c["P1"] = "Release"; ws_c["Q1"] = "Count"
for i, (k, v) in enumerate(release_cnts.items(), 2):
    ws_c.cell(i, 16).value = k; ws_c.cell(i, 17).value = int(v)

# ─────────────────────────────────────────────────────────────────────────────
# ── 2.  DASHBOARD SHEET ────────────────────────────────────────────────────────
# ─────────────────────────────────────────────────────────────────────────────
ws_dash = wb.create_sheet("Dashboard", 0)
ws_dash.sheet_view.showGridLines = False
ws_dash.sheet_view.zoomScale = 85
ws_dash.freeze_panes = "A1"

# Column layout (A=margin, B-G=KPIs/director, H=gap, I-N=right side)
col_w = [1.5, 16, 16, 16, 16, 16, 16, 2.5, 16, 16, 16, 16, 16, 16, 1.5]
for i, w in enumerate(col_w, 1):
    ws_dash.column_dimensions[get_column_letter(i)].width = w

# ── Banner ─────────────────────────────────────────────────────────────────────
for r in range(1, 6):
    ws_dash.row_dimensions[r].height = {1:5, 2:38, 3:5, 4:16, 5:8}[r]
    for c in range(2, 16):
        ws_dash.cell(r, c).fill = fill(NAVY)

ws_dash.merge_cells("B2:O2")
c = ws_dash["B2"]
c.value = "ROCm  Defect Dashboard"
c.font  = Font(bold=True, size=24, color=WHITE, name="Calibri")
c.fill  = fill(NAVY)
c.alignment = Alignment(horizontal="center", vertical="center")

ws_dash.merge_cells("B4:O4")
c = ws_dash["B4"]
c.value = (f"Software Development  |  Defect Tracking & Analytics  |  "
           f"Generated: {datetime.now().strftime('%d %b %Y')}")
c.font  = Font(size=9, color="AAAAAA", italic=True, name="Calibri")
c.fill  = fill(NAVY)
c.alignment = Alignment(horizontal="center", vertical="center")

# ── KPI Cards ──────────────────────────────────────────────────────────────────
kpi_keys   = list(KPI_VALUES.keys())
kpi_vals   = list(KPI_VALUES.values())

for row_h, row in zip([16, 36, 12, 14, 8], [6, 7, 8, 9, 10]):
    ws_dash.row_dimensions[row].height = row_h

for idx, ((title, bg, fg), val) in enumerate(zip(KPI_STYLES, kpi_vals)):
    col = idx + 2   # columns B-G
    for row in [6, 7, 8, 9]:
        ws_dash.cell(row, col).fill = fill(bg)
    # Merge rows 7+8 for big number
    ws_dash.merge_cells(
        start_row=7, start_column=col,
        end_row=8,   end_column=col
    )
    # Label
    c = ws_dash.cell(6, col)
    c.value = title
    c.font  = Font(bold=True, size=7.5, color=fg, name="Calibri")
    c.alignment = Alignment(horizontal="center", vertical="center")
    # Big number
    c = ws_dash.cell(7, col)
    c.value = val
    c.font  = Font(bold=True, size=32, color=fg, name="Calibri")
    c.alignment = Alignment(horizontal="center", vertical="center")
    # Sub-label
    c = ws_dash.cell(9, col)
    c.value = KPI_STYLES[idx][0].title().replace("_"," ")
    c.font  = Font(size=7.5, color=fg, italic=True, name="Calibri")
    c.alignment = Alignment(horizontal="center", vertical="center")

# Sub-label overrides (friendlier text)
sub_labels = ["All defects", "Needs attention", "Highest severity",
              "Priority 1", "Being worked", "Done"]
for idx, sub in enumerate(sub_labels):
    ws_dash.cell(9, idx + 2).value = sub

# ── Director Summary Table ─────────────────────────────────────────────────────
ws_dash.row_dimensions[11].height = 8
ws_dash.row_dimensions[12].height = 16

ws_dash.merge_cells("B12:G12")
c = ws_dash["B12"]
c.value = "DIRECTOR SUMMARY"
c.font  = Font(bold=True, size=10, color=WHITE, name="Calibri")
c.fill  = fill(SLATE)
c.alignment = Alignment(horizontal="left", vertical="center")

# Table headers
tbl_heads = ["Director", "Teams", "Total", "Open", "Critical", "Resolved"]
tbl_bg    = "2C3E50"
for ci, h in enumerate(tbl_heads, 2):
    c = ws_dash.cell(13, ci)
    c.value = h
    c.font  = Font(bold=True, size=9, color=WHITE, name="Calibri")
    c.fill  = fill(tbl_bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
ws_dash.row_dimensions[13].height = 16

# Table rows
for ri_off, director in enumerate(DIRECTOR_ORDER):
    row = 14 + ri_off
    ws_dash.row_dimensions[row].height = 16
    row_bg = LGRAY if ri_off % 2 == 0 else WHITE
    d_total = int(dir_sum.loc[director, "Total"])   if director in dir_sum.index else 0
    d_open  = int(dir_sum.loc[director, "Open"])    if director in dir_sum.index else 0
    d_crit  = int(dir_sum.loc[director, "Critical"])if director in dir_sum.index else 0
    d_res   = int(dir_sum.loc[director, "Resolved"])if director in dir_sum.index else 0

    row_vals = [director, DIRECTOR_TEAMS.get(director, ""), d_total, d_open, d_crit, d_res]
    for ci, val in enumerate(row_vals, 2):
        c = ws_dash.cell(row, ci)
        c.value = val
        c.fill  = fill(row_bg)
        c.alignment = Alignment(horizontal="left" if ci <= 3 else "center",
                                vertical="center")
        c.border = border("CCCCCC")
        if ci == 2:
            c.font = Font(bold=True, size=9, color=DCOL_HEX.get(director, "000000"), name="Calibri")
        elif ci == 4:  # Open column - red if > 0
            c.font = Font(bold=d_open > 0, size=9,
                          color=("C0392B" if d_open > 0 else "000000"), name="Calibri")
        else:
            c.font = Font(size=9, name="Calibri")

ws_dash.row_dimensions[19].height = 8   # spacer before charts

# Set uniform row height (15 pt) for all chart rows to ensure no overlap
# 10 cm chart @ 15 pt/row = ~19 rows tall; 24-row section spacing = 2.7 cm buffer
for _r in range(20, 130):
    ws_dash.row_dimensions[_r].height = 15

# ── Chart Factory Helper ────────────────────────────────────────────────────────
def stacked_bar(ws_src, n_rows, x_col, series_cols, title,
                chart_type="col", width_cm=16, height_cm=10):
    chart = BarChart()
    chart.type      = chart_type
    chart.grouping  = "stacked"
    chart.overlap   = 100
    chart.title     = title
    chart.style     = 10
    chart.width     = width_cm
    chart.height    = height_cm
    chart.legend.position = "b"
    cats = Reference(ws_src, min_col=x_col, min_row=2, max_row=n_rows + 1)
    for col_idx in series_cols:
        data = Reference(ws_src, min_col=col_idx, min_row=1, max_row=n_rows + 1)
        chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    return chart

def pie_chart(ws_src, lbl_col, val_col, n_rows, title, width_cm=8, height_cm=10):
    chart = PieChart()
    chart.title  = title
    chart.style  = 10
    chart.width  = width_cm
    chart.height = height_cm
    labels = Reference(ws_src, min_col=lbl_col, min_row=2, max_row=n_rows + 1)
    data   = Reference(ws_src, min_col=val_col, min_row=1, max_row=n_rows + 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)
    return chart

def bar_chart(ws_src, lbl_col, val_col, n_rows, title,
              chart_type="col", width_cm=16, height_cm=10):
    chart = BarChart()
    chart.type      = chart_type
    chart.grouping  = "clustered"
    chart.title     = title
    chart.style     = 10
    chart.width     = width_cm
    chart.height    = height_cm
    chart.legend    = None
    cats = Reference(ws_src, min_col=lbl_col, min_row=2, max_row=n_rows + 1)
    data = Reference(ws_src, min_col=val_col, min_row=1, max_row=n_rows + 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    return chart

def line_chart(ws_src, n_rows, x_col, series_cols, title, width_cm=16, height_cm=10):
    chart = LineChart()
    chart.title  = title
    chart.style  = 10
    chart.width  = width_cm
    chart.height = height_cm
    chart.legend.position = "b"
    cats = Reference(ws_src, min_col=x_col, min_row=2, max_row=n_rows + 1)
    for col_idx in series_cols:
        data = Reference(ws_src, min_col=col_idx, min_row=1, max_row=n_rows + 1)
        chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    return chart

# ─────────────────────────────────────────────────────────────────────────────
# CHART GRID — 4 rows × 2 columns, uniform 10 cm height, 24-row spacing
#   Each chart 16 cm wide fits in its 6-col section (6 × 2.96 cm ≈ 17.8 cm)
#   Row anchors: 21, 45, 69, 93  (24-row gap = 12.7 cm > 10 cm chart → 2.7 cm buffer)
#   Pies: 8 cm wide → two pies fit side-by-side in the 17.8 cm right section
# ─────────────────────────────────────────────────────────────────────────────

def _label(row, text):
    ws_dash.row_dimensions[row].height = 14
    ws_dash.merge_cells(f"B{row}:N{row}")
    c = ws_dash[f"B{row}"]
    c.value = text
    c.font      = Font(bold=True, size=9, color="AAAAAA", name="Calibri")
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.fill      = fill("F5F6FA")

_label(20, "TREND  &  DIRECTOR OVERVIEW")
_label(44, "TRIAGE  &  WoW BY DIRECTOR")
_label(68, "PRIORITY  /  SEVERITY  /  STATUS  DISTRIBUTION")
_label(92, "TOP CATEGORIES  &  TOP ASSIGNEES")

# ── Row 1  (anchor row 21): WoW stacked bar  |  Director × Severity bar ───────
ch_wow = stacked_bar(ws_wow, n_wow, x_col=1, series_cols=[2, 3, 4, 5],
                     title="Week-on-Week Issues Created  (by Priority)",
                     chart_type="col")
ws_dash.add_chart(ch_wow, "B21")

ch_dir = stacked_bar(ws_dir, n_dir, x_col=1, series_cols=[2, 3, 4, 5],
                     title="Issues by Director  (by Severity)",
                     chart_type="bar")
ws_dash.add_chart(ch_dir, "I21")

# ── Row 2  (anchor row 45): Triage × Status bar  |  WoW by Director line ──────
ch_triage = stacked_bar(ws_tr, n_tr, x_col=1, series_cols=list(range(2, 8)),
                        title="Issues by Triage Assignment  (by Status)",
                        chart_type="bar")
ws_dash.add_chart(ch_triage, "B45")

ch_wdir = line_chart(ws_wdir, n_wdir, x_col=1, series_cols=list(range(2, 7)),
                     title="Week-on-Week Issues by Director")
ws_dash.add_chart(ch_wdir, "I45")

# ── Row 3  (anchor row 69): Status bar  |  Priority pie  |  Severity pie ───────
ch_status = bar_chart(ws_c, lbl_col=7, val_col=8, n_rows=len(STATUS_ORDER),
                      title="Status Breakdown", chart_type="col")
ws_dash.add_chart(ch_status, "B69")

ch_ppie = pie_chart(ws_c, lbl_col=1, val_col=2, n_rows=len(PRIORITY_ORDER),
                    title="Issues by Priority")
ws_dash.add_chart(ch_ppie, "I69")

ch_spie = pie_chart(ws_c, lbl_col=4, val_col=5, n_rows=len(SEVERITY_ORDER),
                    title="Issues by Severity")
ws_dash.add_chart(ch_spie, "L69")

# ── Row 4  (anchor row 93): Top Categories bar  |  Top Assignees bar ──────────
ch_cat = bar_chart(ws_c, lbl_col=10, val_col=11, n_rows=len(category_cnts),
                   title="Top Triage Categories", chart_type="bar")
ws_dash.add_chart(ch_cat, "B93")

ch_asgn = bar_chart(ws_c, lbl_col=13, val_col=14, n_rows=len(top_asgn),
                    title="Top Assignees  (Open Issues)", chart_type="bar")
ws_dash.add_chart(ch_asgn, "I93")

# ─────────────────────────────────────────────────────────────────────────────
# ── 3.  DEFECTS DATA SHEET ─────────────────────────────────────────────────────
# ─────────────────────────────────────────────────────────────────────────────
ws_data = wb.create_sheet("Defects")
ws_data.freeze_panes = "A2"

DATA_COLS = [
    "Defect_ID", "Title", "Created_Date", "Created_By",
    "Priority", "Severity", "Status",
    "Triage_Category", "Triage_Assignment", "Director", "Assignee",
    "Component", "Target_Release", "Found_In_Release",
    "Last_Updated_Date", "Resolution",
]
COL_WIDTHS = [12, 48, 13, 18, 10, 15, 13, 28, 24, 10, 18, 18, 14, 14, 16, 14]
for ci, w in enumerate(COL_WIDTHS, 1):
    ws_data.column_dimensions[get_column_letter(ci)].width = w

# Header row
ws_data.row_dimensions[1].height = 20
for ci, col in enumerate(DATA_COLS, 1):
    c = ws_data.cell(1, ci)
    c.value     = col.replace("_", " ")
    c.font      = Font(bold=True, size=10, color=WHITE, name="Calibri")
    c.fill      = fill(NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border    = border(WHITE)

# Conditional fills
P_FILL  = {"Critical":"FADBD8","High":"FDEBD0","Medium":"FEF9E7","Low":"EAFAF1"}
S_FILL  = {"S1 - Blocker":"FADBD8","S2 - Critical":"FDEBD0",
           "S3 - Major":"FEF9E7","S4 - Minor":"EAFAF1"}
ST_FILL = {"Open":"FADBD8","In Progress":"FDEBD0","In Review":"FEF9E7",
           "Resolved":"EAFAF1","Closed":"ECF0F1","Won't Fix":"F4ECF7"}

for ri, (_, row) in enumerate(df[DATA_COLS].iterrows(), 2):
    ws_data.row_dimensions[ri].height = 15
    row_bg = "FAFAFA" if ri % 2 == 0 else WHITE
    for ci, col in enumerate(DATA_COLS, 1):
        val = row[col]
        if col in ("Created_Date", "Last_Updated_Date") and pd.notna(val):
            val = pd.to_datetime(val).strftime("%Y-%m-%d")
        c = ws_data.cell(ri, ci)
        c.value     = "" if pd.isna(val) else val
        c.alignment = Alignment(
            horizontal="left" if ci == 2 else "center",
            vertical="center", wrap_text=(ci == 2)
        )
        c.border = border()
        if   col == "Priority" and val in P_FILL:
            c.fill = fill(P_FILL[val])
            c.font = Font(bold=True, size=9, name="Calibri")
        elif col == "Severity" and val in S_FILL:
            c.fill = fill(S_FILL[val])
            c.font = Font(size=9, name="Calibri")
        elif col == "Status" and val in ST_FILL:
            c.fill = fill(ST_FILL[val])
            c.font = Font(size=9, name="Calibri")
        elif col == "Director" and val in DCOL_HEX:
            c.fill = fill(row_bg)
            c.font = Font(bold=True, size=9, color=DCOL_HEX[val], name="Calibri")
        else:
            c.fill = fill(row_bg)
            c.font = Font(size=9, name="Calibri")

# AutoFilter
ws_data.auto_filter.ref = f"A1:{get_column_letter(len(DATA_COLS))}{len(df)+1}"

# ─────────────────────────────────────────────────────────────────────────────
# ── 4.  SUMMARY PIVOT SHEET ────────────────────────────────────────────────────
# ─────────────────────────────────────────────────────────────────────────────
ws_piv = wb.create_sheet("Summary")
ws_piv.sheet_view.showGridLines = False

def piv_title(ws, row, col, text, span=4):
    ws.merge_cells(start_row=row, start_column=col,
                   end_row=row,   end_column=col + span - 1)
    c = ws.cell(row, col)
    c.value = text
    c.font  = Font(bold=True, size=10, color=WHITE, name="Calibri")
    c.fill  = fill(SLATE)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 16

def piv_table(ws, start_row, start_col, headers, rows, col_w=None):
    for ci, h in enumerate(headers, start_col):
        c = ws.cell(start_row, ci)
        c.value = h
        c.font  = Font(bold=True, size=9, color=WHITE, name="Calibri")
        c.fill  = fill("344055")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border(WHITE)
    for ri, row_vals in enumerate(rows, start_row + 1):
        row_bg = LGRAY if ri % 2 == 0 else WHITE
        for ci, val in enumerate(row_vals, start_col):
            c = ws.cell(ri, ci)
            c.value = int(val) if isinstance(val, (int, float)) else val
            c.font  = Font(size=9, name="Calibri")
            c.fill  = fill(row_bg)
            c.alignment = Alignment(horizontal="center" if ci > start_col else "left",
                                    vertical="center")
            c.border = border()

# Col widths
for i in range(1, 20):
    ws_piv.column_dimensions[get_column_letter(i)].width = 16

# ── Priority breakdown
piv_title(ws_piv, 1, 1, "Issues by Priority", 2)
piv_table(ws_piv, 2, 1, ["Priority", "Count"],
          [[k, int(v)] for k, v in priority_cnts.items()])

# ── Severity breakdown
piv_title(ws_piv, 1, 4, "Issues by Severity", 2)
piv_table(ws_piv, 2, 4, ["Severity", "Count"],
          [[k, int(v)] for k, v in severity_cnts.items()])

# ── Status breakdown
piv_title(ws_piv, 1, 7, "Issues by Status", 2)
piv_table(ws_piv, 2, 7, ["Status", "Count"],
          [[k, int(v)] for k, v in status_cnts.items()])

# ── Director breakdown (full)
piv_title(ws_piv, 10, 1, "Director Breakdown", 6)
dir_piv_headers = ["Director", "Total", "Open", "Critical", "In Progress", "Resolved"]
dir_piv_rows = []
for d in DIRECTOR_ORDER:
    in_prog = int(df[(df["Director"]==d) & (df["Status"]=="In Progress")].shape[0])
    dir_piv_rows.append([
        d,
        int(dir_sum.loc[d,"Total"])    if d in dir_sum.index else 0,
        int(dir_sum.loc[d,"Open"])     if d in dir_sum.index else 0,
        int(dir_sum.loc[d,"Critical"]) if d in dir_sum.index else 0,
        in_prog,
        int(dir_sum.loc[d,"Resolved"]) if d in dir_sum.index else 0,
    ])
piv_table(ws_piv, 11, 1, dir_piv_headers, dir_piv_rows)

# ── Triage category breakdown
piv_title(ws_piv, 10, 8, "Issues by Triage Category", 2)
piv_table(ws_piv, 11, 8, ["Triage Category", "Count"],
          [[k, int(v)] for k, v in category_cnts.items()])

# ── Week-on-Week summary (most recent 8 weeks)
piv_title(ws_piv, 20, 1, "Recent 8-Week Trend (by Priority)", 5)
wow_recent = wow_raw.tail(8)[["Week_Label"] + PRIORITY_ORDER]
piv_table(ws_piv, 21, 1, ["Week"] + PRIORITY_ORDER,
          wow_recent.values.tolist())

# ─────────────────────────────────────────────────────────────────────────────
# ── Save ───────────────────────────────────────────────────────────────────────
# ─────────────────────────────────────────────────────────────────────────────
out = "ROCm_Defect_Dashboard.xlsx"
wb.save(out)
print(f"Saved: {out}")
print(f"  Sheets: Dashboard (charts + KPIs), Defects ({len(df)} rows), Summary (pivot tables)")
print(f"  Charts: WoW, Director-Severity, Triage-Status, Priority pie,")
print(f"          Severity pie, Director WoW line, Status, Top Categories, Top Assignees")
